import csv
from plistlib import InvalidFileException
from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
import openpyxl
from textblob import TextBlob
# Create your views here.

MAX_INDEX = 4


class Adjectives(APIView):
    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file found in request'}, status=400)
        file = request.FILES['file']
        if 'index' not in request.data:
            return Response({'error': 'No index found in request'}, status=400)
        try:
            index = int(request.data['index'])
        except:
            return Response({'error': 'Index must be an integer'}, status=400)
        # Read the Excel file from the request
        file = request.FILES['file']
        index = int(request.data['index'])
        if index > MAX_INDEX:
            return Response('Invalid Index', status=400)
        if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(file)
            except InvalidFileException:
                return Response('Invalid Excel file', status=400)

            # Get the first sheet
            worksheet = workbook.worksheets[0]

            word_counts = {}

            for row in worksheet.iter_rows():
                if row[index].value:
                    blob = TextBlob(row[index].value)
                    for (word, tag) in blob.tags:
                        if tag == "JJ":
                            # if the word is not in the dictionary, add it with a count of 1
                            if word not in word_counts:
                                word_counts[word] = 1
                            # if the word is already in the dictionary, increment its count
                            else:
                                word_counts[word] += 1

            word_counts = {k: v for k, v in sorted(
                word_counts.items(), key=lambda x: x[1], reverse=True)}
            # Return the data as JSON
            return Response(word_counts)

        elif file.content_type == 'text/csv':
            data = []
            reader = csv.reader(file)
            word_counts = {}
            for row in reader:
                if row[index].value:
                    blob = TextBlob(row[index].value)
                    for (word, tag) in blob.tags:
                        if tag == "JJ":
                            # if the word is not in the dictionary, add it with a count of 1
                            if word not in word_counts:
                                word_counts[word] = 1
                            # if the word is already in the dictionary, increment its count
                            else:
                                word_counts[word] += 1

            word_counts = {k: v for k, v in sorted(
                word_counts.items(), key=lambda x: x[1], reverse=True)}
            # Return the data as JSON
            return Response(word_counts)
        else:
            return Response('Unsupported file type', status=400)


class Categorize(APIView):

    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file found in request'}, status=400)
        file = request.FILES['file']
        if 'index' not in request.data:
            return Response({'error': 'No index found in request'}, status=400)
        try:
            index = int(request.data['index'])
        except:
            return Response({'error': 'Index must be an integer'}, status=400)
        if index > MAX_INDEX:
            return Response('Invalid Index', status=400)
        if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(file)
            except InvalidFileException:
                return Response('Invalid Excel file', status=400)

            # Get the first sheet
            worksheet = workbook.worksheets[0]
            categories_count = {}

            for row in worksheet.iter_rows():
                if row[index].value:
                    blob = row[index].value
                    # if the word is not in the dictionary, add it with a count of 1
                    if blob not in categories_count:
                        categories_count[blob] = 1
                    # if the word is already in the dictionary, increment its count
                    else:
                        categories_count[blob] += 1

            categories_count = {k: v for k, v in sorted(
                categories_count.items(), key=lambda x: x[1], reverse=True)}

            return Response(categories_count)

        elif file.content_type == 'text/csv':

            reader = csv.reader(file)

            categories_count = {}

            for row in reader:
                if row[index].value:
                    blob = row[index].value
                    # if the word is not in the dictionary, add it with a count of 1
                    if blob not in categories_count:
                        categories_count[blob] = 1
                    # if the word is already in the dictionary, increment its count
                    else:
                        categories_count[blob] += 1

            categories_count = {k: v for k, v in sorted(
                categories_count.items(), key=lambda x: x[1], reverse=True)}

            return Response(categories_count)
        else:
            return Response('Unsupported file type', status=400)


class Nouns(APIView):
    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file found in request'}, status=400)
        file = request.FILES['file']
        if 'index' not in request.data:
            return Response({'error': 'No index found in request'}, status=400)
        try:
            index = int(request.data['index'])
        except:
            return Response({'error': 'Index must be an integer'}, status=400)
        file = request.FILES['file']
        index = int(request.data['index'])
        if index > MAX_INDEX:
            return Response('Invalid Index', status=400)
        if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(file)
            except InvalidFileException:
                return Response('Invalid Excel file', status=400)

            # Get the first sheet
            worksheet = workbook.worksheets[0]

            word_counts = {}

            for row in worksheet.iter_rows():
                if row[index].value:
                    blob = TextBlob(row[index].value)
                    for (word, tag) in blob.tags:
                        if tag == "NN":
                            # if the word is not in the dictionary, add it with a count of 1
                            if word not in word_counts:
                                word_counts[word] = 1
                            # if the word is already in the dictionary, increment its count
                            else:
                                word_counts[word] += 1

            word_counts = {k: v for k, v in sorted(
                word_counts.items(), key=lambda x: x[1], reverse=True)}
            # Return the data as JSON
            return Response(word_counts)

        elif file.content_type == 'text/csv':
            data = []
            reader = csv.reader(file)
            word_counts = {}
            for row in reader:
                if row[index].value:
                    blob = TextBlob(row[index].value)
                    for (word, tag) in blob.tags:
                        if tag == "NN":
                            # if the word is not in the dictionary, add it with a count of 1
                            if word not in word_counts:
                                word_counts[word] = 1
                            # if the word is already in the dictionary, increment its count
                            else:
                                word_counts[word] += 1

            word_counts = {k: v for k, v in sorted(
                word_counts.items(), key=lambda x: x[1], reverse=True)}
            # Return the data as JSON
            return Response(word_counts)
        else:
            return Response('Unsupported file type', status=400)


class RankNegativeSentiments(APIView):
    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file found in request'}, status=400)
        file = request.FILES['file']
        if 'index' not in request.data:
            return Response({'error': 'No index found in request'}, status=400)
        try:
            index = int(request.data['index'])
        except:
            return Response({'error': 'Index must be an integer'}, status=400)
        # Read the Excel file from the request
        file = request.FILES['file']
        index = int(request.data['index'])
        if index > MAX_INDEX:
            return Response('Invalid Index', status=400)
        if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(file)
            except InvalidFileException:
                return Response('Invalid Excel file', status=400)

            # Get the first sheet
            worksheet = workbook.worksheets[0]
            sentiments = []

            for row in worksheet.iter_rows():
                if row[index].value:
                    testimonial = TextBlob(row[index].value)
                    result = testimonial.sentiment
                    sentiments.append({result.polarity: f"{row[index].value}"})

            # sort positive to negative
            sentiments = sorted(sentiments, key=lambda x: list(
                x.keys())[0])

            # Return the data as JSON
            return Response(sentiments)

        elif file.content_type == 'text/csv':

            reader = csv.reader(file)
            sentiments = []
            for row in reader:
                if row[index].value:
                    testimonial = TextBlob(row[index].value)
                    result = testimonial.sentiment
                    sentiments.append({result.polarity: f"{row[index].value}"})

            # sort negative to positive
            sentiments = sorted(sentiments, key=lambda x: list(
                x.keys())[0])

            # Return the data as JSON
            return Response(sentiments)
        else:
            return Response('Unsupported file type', status=400)


class RankPositiveSentiments(APIView):
    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file found in request'}, status=400)
        file = request.FILES['file']
        if 'index' not in request.data:
            return Response({'error': 'No index found in request'}, status=400)
        try:
            index = int(request.data['index'])
        except:
            return Response({'error': 'Index must be an integer'}, status=400)
        # Read the Excel file from the request
        file = request.FILES['file']
        index = int(request.data['index'])
        if index > MAX_INDEX:
            return Response('Invalid Index', status=400)
        if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(file)
            except InvalidFileException:
                return Response('Invalid Excel file', status=400)

            # Get the first sheet
            worksheet = workbook.worksheets[0]
            sentiments = []

            for row in worksheet.iter_rows():
                if row[index].value:
                    testimonial = TextBlob(row[index].value)
                    result = testimonial.sentiment
                    sentiments.append({result.polarity: f"{row[index].value}"})

            # sort positive to negative
            sentiments = sorted(sentiments, key=lambda x: list(
                x.keys())[0], reverse=True)

            # Return the data as JSON
            return Response(sentiments)

        elif file.content_type == 'text/csv':

            reader = csv.reader(file)
            sentiments = []
            for row in reader:
                if row[index].value:
                    testimonial = TextBlob(row[index].value)
                    result = testimonial.sentiment
                    sentiments.append({result.polarity: f"{row[index].value}"})

            # sort positive to negative
            sentiments = sorted(sentiments, key=lambda x: list(
                x.keys())[0], reverse=True)

            # Return the data as JSON
            return Response(sentiments)
        else:
            return Response('Unsupported file type', status=400)
